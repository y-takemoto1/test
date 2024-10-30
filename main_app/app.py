# coding:utf-8

# 必要なパッケージのインポート
import streamlit as st
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome import service as fs
from selenium.webdriver import ChromeOptions
from webdriver_manager.core.os_manager import ChromeType
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl

# タイトルを設定
st.title("seleniumテストアプリ")

# ボタンを作成(このボタンをアプリ上で押すと"if press_button:"より下の部分が実行される)
press_button = st.button("スクレイピング開始")

# エクセルを開く
wb = openpyxl.Workbook()
ws = wb.active

if press_button:
    # スクレイピングするwebサイトのURL
    URL = 'https://jp.indeed.com/jobs?q=%E6%AD%A3%E7%A4%BE%E5%93%A1&l=%E7%A6%8F%E5%B2%A1%E7%9C%8C&from=searchOnDesktopSerp&vjk=d4bb56841be7498e'

    # ドライバのオプション
    options = ChromeOptions()
    # option設定を追加（設定する理由はメモリの削減）
    options.add_argument("--headless")
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    # webdriver_managerによりドライバーをインストール
    CHROMEDRIVER = ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()
    service = fs.Service(CHROMEDRIVER)
    driver = webdriver.Chrome(
                              options=options,
                              service=service
                             )

    # URLで指定したwebページを開く
    driver.get(URL)
    # ページが読み込まれるまで待機
    wait = WebDriverWait(driver, 20)
    st.text('ルート確認１')
    c = 1
    j = 1
    k = 1
    # 1ページにつき15件あるため15件単位取得される(例：10にすると150件取得可能)
    #　 　　　　↓
    while j <= 10:
        print('COUNT:', c)
        time.sleep(5)
        st.text('ルート確認2')
        while k <= 2:
            st.text('ルート1')
            elements = driver.find_elements(By.XPATH, '//*')
            st.text(elements)
            # 各要素のテキストや属性を表示
            for element in elements:
                st.text(element.text)  # テキストを取得
                # 例えば、特定の属性の値を取得したい場合

            st.text('ルート2')
            try:
                links = wait.until(EC.presence_of_all_elements_located((By.ID, 'mainContentTable')))
                #links = driver.find_element(By.CLASS_NAME, 'job_seen_beacon')
                st.text('ルート確認3')
                st.text(f'LINKSの中身：{links}')
                break
            except Exception as e:
                k += 1
                st.text('ルート確認4')
                st.text(f"求人情報の取得に失敗: {e}")
                driver.quit()
                st.stop()  # エラーが発生したら処理を終了
        print('ルート確認11111')

        for link in links:
            # リンクをクリックする前に、再度要素を取得
            try:
                wait.until(EC.element_to_be_clickable(link)).click()
            except Exception as e:
                print("リンクのクリックに失敗:", e)
                continue  # 失敗した場合は次のリンクに進む

            st.text('ルート確認5')
            # 詳細ページが読み込まれるまで待機
            wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'jobsearch-HeaderContainer')))
            time.sleep(10)
            st.text('ルート確認6')

            # 初期値を設定
            title = location = company = tel = desired_element = "該当無し"

            # 特定の要素を取得
            try:
                custom_data_element = driver.find_elements(By.CLASS_NAME, 'css-1jkk9vb')
                if custom_data_element:
                    desired_element = custom_data_element[0].text
                    print('仕事内容：', desired_element)
                print('-' * 80)

                try:
                    title = driver.find_element(By.CLASS_NAME, 'jobsearch-JobInfoHeader-title').text
                    print('title:', title)
                except Exception as e:
                    print('タイトルの取得に失敗:')            
                print('-' * 80)

                try:
                    location = driver.find_element(By.CLASS_NAME, 'css-scnhev').text
                    print('location:', location)
                except Exception as e:
                    print('勤務地の取得に失敗:')
                print('-' * 80)

                try:
                    company = driver.find_element(By.CLASS_NAME, 'css-1f8zkg3').text
                    print('company:', company)
                except Exception as e:
                    print('会社名の取得に失敗:')
                print('-' * 80)

                # 特定の文字列を持つ要素を検索
                try:
                    parent_element = driver.find_element(By.XPATH, "//div[text()='代表電話番号']")
                    child_elements = parent_element.find_elements(By.XPATH, "../*")
                    tel = child_elements[1].text if len(child_elements) > 1 else "該当無し"
                    print('電話番号：', child_elements[1].text if len(child_elements) > 1 else "該当なし")
                except Exception as e:
                    print('電話番号の取得に失敗:')
                print('-' *80)
            except Exception as e:
                print('詳細情報の取得に失敗:')
                print('-' *80)
            
            ws.cell(row=c, column=1, value=title)
            ws.cell(row=c, column=2, value=location)
            ws.cell(row=c, column=3, value=company)
            ws.cell(row=c, column=4, value=tel)
            ws.cell(row=c, column=5, value=desired_element)
            
            c += 1
            st.text(f'抽出件数：{c}件目')
            # 詳細ページから戻る
            #driver.back()
            wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'job_seen_beacon')))
            time.sleep(10)
        # スクロール
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        # 次のページを探す
        try:
            next_page = wait.until(EC.presence_of_element_located((By.XPATH, "//a[@aria-label='Next Page']")))
            print('NEXT:', next_page.text)
            if next_page:
                next_page.click()
        except Exception as e:
            print('次のページの取得に失敗:')
            break  # 次のページが取得できなかったらループを終了
        # ポップアップが表示されている場合は閉じる
        try:
            # 一定時間待機してからポップアップをチェック
            wait.until(EC.presence_of_element_located((By.XPATH, "//button[contains(@class, 'css-yi9ndv')]")))
            close_button = driver.find_element(By.XPATH, "//button[contains(@class, 'css-yi9ndv')]")
            time.sleep(10)
            if close_button.is_displayed():  # ボタンが表示されているか確認
                close_button.click()
                print('ポップアップを閉じました。')
        except Exception as e:
            print('ポップアップ無し') 
        j += 1

    wb.save("INDEED.xlsx")
    wb.close()

    time.sleep(10)

    # webページを閉じる
    driver.close()

    # スクレピン完了したことをstreamlitアプリ上に表示する
    st.write("スクレイピング完了!!!")